from tkinter import *
from tkinter import ttk
from PIL import Image, ImageTk
from tkinter import messagebox
from openpyxl import load_workbook
import cv2
import os

class RegistrationPage:
    def __init__(self, root):
        self.root = root
        self.root.geometry("1530x790+0+0")
        self.root.title("Registration")

        self.root = root
        self.dep_combo = None
        self.course_combo = None
        self.year_combo = None
        self.semester_combo = None
        self.section_combo = None
        self.student_Id_entry = None
        self.student_Name_entry = None
        self.phone_entry = None
        self.email_entry = None
        self.student_table = None
        self.student_data = []

        self.scholar_id = ""
        self.cap = None
        self.photo_taken = False

    

        # top image
        img = Image.open(r"Images\cover.jpeg")
        img = img.resize((1530, 130))
        self.top = ImageTk.PhotoImage(img)

        top = Label(self.root, image=self.top)
        top.place(x=0, y=0, width=1530, height=130)

        # Title
        title_lbl = Label(self.root, text="STUDENT MANAGEMENT SYSTEM", font=("times new roman", 35, "bold"), bg="white", fg="black")
        title_lbl.place(x=0, y=130, width=1530, height=45)

        # main frame
        main_frame = Frame(self.root, bd=2, bg="white")
        main_frame.place(x=20, y=180, width=1480, height=600)

        # left label frame
        Left_frame = LabelFrame(main_frame, bd=2, bg="white", relief=RIDGE, text="Student Details", font=("times new roman", 12, "bold"))
        Left_frame.place(x=10, y=10, width=730, height=580)

        img_left = Image.open(r"Images\20210121collegestudents.jpg")
        img_left = img_left.resize((720, 130))
        self.photoimg_left = ImageTk.PhotoImage(img_left)

        img_left_lbl = Label(Left_frame, image=self.photoimg_left)
        img_left_lbl.place(x=5, y=0, width=720, height=130)

        # current Course
        current_course_frame = LabelFrame(Left_frame, bd=2, bg="white", relief=RIDGE, text="Current Course", font=("times new roman", 12, "bold"))
        current_course_frame.place(x=5, y=140, width=720, height=160)

        # Department
        dep_lbl = Label(current_course_frame, text="Department", font=("times new roman", 12, "bold"), bg="white")
        dep_lbl.grid(row=0, column=0, padx=10, sticky=W)

        dep_combo = ttk.Combobox(current_course_frame, font=("times new roman", 10, "bold"), state="readonly", width=17)
        dep_combo["values"] = ("Select Department", "CSE", "ECE", "Electrical", "Mechanical", "MDS", "Chemical", "MME")
        dep_combo.current(0)
        dep_combo.grid(row=0, column=1, padx=2, pady=10, sticky=W)
        self.dep_combo = dep_combo

        # Course
        course_lbl = Label(current_course_frame, text="Course", font=("times new roman", 12, "bold"), bg="white")
        course_lbl.grid(row=0, column=2, padx=10, sticky=W)

        course_combo = ttk.Combobox(current_course_frame, font=("times new roman", 10, "bold"), state="readonly", width=17)
        course_combo["values"] = ("Select Course", "B.Tech", "M.Tech", "PHD")
        course_combo.current(0)
        course_combo.grid(row=0, column=3, padx=2, pady=10, sticky=W)
        self.course_combo = course_combo

        # Year
        year_lbl = Label(current_course_frame, text="Year", font=("times new roman", 12, "bold"), bg="white")
        year_lbl.grid(row=1, column=0, padx=10, sticky=W)
        year_combo = ttk.Combobox(current_course_frame, font=("times new roman", 10, "bold"), state="readonly", width=17)
        year_combo["values"] = ("Select Year", "1st Year", "2nd Year", "3rd Year", "4th Year")
        year_combo.current(0)
        year_combo.grid(row=1, column=1, padx=2, pady=10, sticky=W)
        self.year_combo = year_combo

        # Semester
        semester_lbl = Label(current_course_frame, text="Semester", font=("times new roman", 12, "bold"), bg="white")
        semester_lbl.grid(row=1, column=2, padx=10, sticky=W)
        semester_combo = ttk.Combobox(current_course_frame, font=("times new roman", 10, "bold"), state="readonly", width=17)
        semester_combo["values"] = ("Select Semester", "1st Semester", "2nd Semester", "3rd Semester", "4th Semester", "5th Semester", "6th Semester", "7th Semester", "8th Semester")
        semester_combo.current(0)
        semester_combo.grid(row=1, column=3, padx=2, pady=10, sticky=W)
        self.semester_combo = semester_combo

        # section
        section_lbl = Label(current_course_frame, text="Section", font=("times new roman", 12, "bold"), bg="white")
        section_lbl.grid(row=2, column=0, padx=10, pady=5, sticky=W)
        section_combo = ttk.Combobox(current_course_frame, font=("times new roman", 10, "bold"), state="readonly", width=17)
        section_combo["values"] = ("Select Section", "I", "II", "III")
        section_combo.current(0)
        section_combo.grid(row=2, column=1, padx=2, pady=5, sticky=W)
        self.section_combo = section_combo

        # Class Student Information

        class_Student_frame = LabelFrame(Left_frame, bd=2, bg="white", relief=RIDGE, text="Class Student Information", font=("times new roman", 12, "bold"))
        class_Student_frame.place(x=5, y=300, width=720, height=255)

        student_Id_lbl = Label(class_Student_frame, text="Student ID", font=("times new roman", 12, "bold"), bg="white")
        student_Id_lbl.grid(row=0, column=0, padx=10, pady=5, sticky=W)
        self.student_Id_entry = ttk.Entry(class_Student_frame, width=20, font=("times new roman", 10, "bold"))
        self.student_Id_entry.grid(row=0, column=1, padx=10, pady=5, sticky=W)
 
        student_Name_lbl = Label(class_Student_frame, text="Student Name", font=("times new roman", 12, "bold"), bg="white")
        student_Name_lbl.grid(row=0, column=2, padx=10, pady=5, sticky=W)
        self.student_Name_entry = ttk.Entry(class_Student_frame, width=20, font=("times new roman", 10, "bold"))
        self.student_Name_entry.grid(row=0, column=3, padx=10, pady=5, sticky=W)

        phone_lbl = Label(class_Student_frame, text="Phone No:", font=("times new roman", 12, "bold"), bg="white")
        phone_lbl.grid(row=1, column=0, padx=10, pady=10, sticky=W)
        self.phone_entry = ttk.Entry(class_Student_frame, width=20, font=("times new roman", 10, "bold"))
        self.phone_entry.grid(row=1, column=1, padx=10, pady=10, sticky=W)

        email_lbl = Label(class_Student_frame, text="Email Id:", font=("times new roman", 12, "bold"), bg="white")
        email_lbl.grid(row=1, column=2, padx=10, pady=10, sticky=W)
        self.email_entry = ttk.Entry(class_Student_frame, width=20, font=("times new roman", 10, "bold"))
        self.email_entry.grid(row=1, column=3, padx=10, pady=10, sticky=W)

        # radio button
        self.radio_var1 = StringVar()
        self.radio_var2 = StringVar()

        # Radiobuttons
        self.radiobtn1 = ttk.Radiobutton(class_Student_frame, text="Take Photo Sample", value="Yes", variable=self.radio_var1)
        self.radiobtn1.grid(row=6, column=0)

        self.radiobtn2 = ttk.Radiobutton(class_Student_frame, text="No Photo Sample", value="No", variable=self.radio_var2)
        self.radiobtn2.grid(row=6, column=1)

        # buttonsFrame
        btn_frame = Frame(class_Student_frame, bd=2, relief=RIDGE, bg="white")
        btn_frame.place(x=0, y=125, width=715, height=35)

        save_btn = Button(btn_frame, text="Save", width=17, font=("times new roman", 13, "bold"), bg="blue", fg="white", command=self.save_data)
        save_btn.grid(row=0, column=0)

        update_btn = Button(btn_frame, text="Update", width=17, font=("times new roman", 13, "bold"), bg="blue", fg="white", command=self.update_data)
        update_btn.grid(row=0, column=1)

        delete_btn = Button(btn_frame, text="Delete", width=17, font=("times new roman", 13, "bold"), bg="blue", fg="white", command=self.delete_data)
        delete_btn.grid(row=0, column=2)

        reset_btn = Button(btn_frame, text="Reset", width=17, font=("times new roman", 13, "bold"), bg="blue", fg="white", command=self.reset_fields)
        reset_btn.grid(row=0, column=3)

        # new frame
        btn_frame1 = Frame(class_Student_frame, bd=2, relief=RIDGE, bg="white")
        btn_frame1.place(x=0, y=160, width=715, height=35)

        takePhoto_btn = Button(btn_frame1, text="Take Photo Sample", width=35, font=("times new roman", 13, "bold"), bg="blue", fg="white", command=self.take_photo)
        takePhoto_btn.grid(row=0, column=0)

        update_photo_btn = Button(btn_frame1, text="Update Photo Sample", width=35, font=("times new roman", 13, "bold"), bg="blue", fg="white", command=self.update_photo)
        update_photo_btn.grid(row=0, column=1)

        # Right label frame
        Right_label_frame = LabelFrame(main_frame, bd=2, bg="white", relief=RIDGE, text="Student Details", font=("times new roman", 12, "bold"))
        Right_label_frame.place(x=750, y=10, width=730, height=580)

        img_right = Image.open(r"Images\Classroom_StudentGroup_Studying_Indoor_GettyImages-679437550.webp")
        img_right = img_right.resize((720, 130))
        self.photoimg_right = ImageTk.PhotoImage(img_right)

        img_right_lbl = Label(Right_label_frame, image=self.photoimg_right)
        img_right_lbl.place(x=5, y=0, width=720, height=130)

        # Search System
        Searching_frame = LabelFrame(Right_label_frame, bd=2, bg="white", relief=RIDGE, text="Search System", font=("times new roman", 12, "bold"))
        Searching_frame.place(x=5, y=135, width=720, height=70)
        search_lbl = Label(Searching_frame, text="Search By:", font=("times new roman", 15, "bold"), bg="red", fg="white")
        search_lbl.grid(row=0, column=0, padx=10, pady=5, sticky=W)

        search_combo = ttk.Combobox(Searching_frame, font=("times new roman", 12, "bold"), state="readonly", width=15)
        search_combo["values"] = ("Select", "Student ID", "Student Name")
        search_combo.current(0)
        search_combo.grid(row=0, column=1, padx=2, pady=10, sticky=W)
        self.search_combo = search_combo

        search_entry = ttk.Entry(Searching_frame, width=15, font=("times new roman", 12, "bold"))
        search_entry.grid(row=0, column=2, padx=10, pady=5, sticky=W)
        self.search_entry = search_entry

        search_btn = Button(Searching_frame, text="Search", width=12, font=("times new roman", 12, "bold"), bg="blue", fg="white", command=self.search_data)
        search_btn.grid(row=0, column=3, padx=10, pady=5, sticky=W)

        showAll_btn = Button(Searching_frame, text="Show All", width=12, font=("times new roman", 12, "bold"), bg="blue", fg="white", command=self.show_data)
        showAll_btn.grid(row=0, column=4, padx=10, pady=5, sticky=W)

        # Table Frame
        table_frame = Frame(Right_label_frame, bd=2, bg="white", relief=RIDGE)
        table_frame.place(x=5, y=210, width=710, height=350)

        scroll_x = ttk.Scrollbar(table_frame, orient=HORIZONTAL)
        scroll_y = ttk.Scrollbar(table_frame, orient=VERTICAL)

        self.student_table = ttk.Treeview(table_frame, column=("dep", "course", "year", "sem","section", "id","name", "email", "phone", "photo"), xscrollcommand=scroll_x.set, yscrollcommand=scroll_y.set)

        scroll_x.pack(side=BOTTOM, fill=X)
        scroll_y.pack(side=RIGHT, fill=Y)

        scroll_x.config(command=self.student_table.xview)
        scroll_y.config(command=self.student_table.yview)
    
        self.student_table.heading("dep", text="Department")
        self.student_table.heading("course", text="Course")
        self.student_table.heading("year", text="Year")
        self.student_table.heading("sem", text="Semester")
        self.student_table.heading("section",text = "Section")
        self.student_table.heading("id", text="Student ID")
        self.student_table.heading("name", text="Student Name")
        self.student_table.heading("email", text="Phone")
        self.student_table.heading("phone", text="Email")
        self.student_table.heading("photo", text="Photo Sample")

        self.student_table["show"] = "headings"

        self.student_table.column("dep", width=100)
        self.student_table.column("course", width=100)
        self.student_table.column("year", width=100)
        self.student_table.column("sem", width=100)
        self.student_table.column("section", width=100)
        self.student_table.column("id", width=100)
        self.student_table.column("name", width=100)
        self.student_table.column("phone", width=100)
        self.student_table.column("email", width=100)

        self.student_table.pack(fill=BOTH, expand=1)
        
    def save_data(self):
        department = self.dep_combo.get()
        course = self.course_combo.get()
        year = self.year_combo.get()
        semester = self.semester_combo.get()
        section = self.section_combo.get()
        student_id = self.student_Id_entry.get()
        student_name = self.student_Name_entry.get()
        phone = self.phone_entry.get()
        email = self.email_entry.get()
        photo_sample = "Yes" if self.radio_var1.get() == "Yes" else "No"  # Capture radio button value

        # Append data to student data list
        self.student_data.append([department, course, year, semester, section, student_id, student_name, phone, email, photo_sample])

        # Save data to Excel file
        wb = load_workbook("student_data.xlsx")
        ws = wb.active
        ws.append([department, course, year, semester, section, student_id, student_name, phone, email,photo_sample])
        wb.save("student_data.xlsx")
        messagebox.showinfo("info","Data saved successfully!")        

    def delete_data(self):
        selected_item = self.student_table.selection()
        if not selected_item:
            messagebox.showerror("Error", "Please select a record to delete.")
            return

        student_id = self.student_table.item(selected_item, 'values')[5]

        wb = load_workbook("student_data.xlsx")
        ws = wb.active

        for row in ws.iter_rows():
            if row[5].value == student_id:
                ws.delete_rows(row[0].row)
                break

        wb.save("student_data.xlsx")
        wb.close()
        messagebox.showinfo("Success", "Record deleted successfully.")
        self.show_data()  # Refresh the displayed data

    def update_data(self):
        selected_item = self.student_table.selection()
        if not selected_item:
            messagebox.showerror("Error", "Please select a record to update.")
            return

        updated_data = [self.dep_combo.get(), self.course_combo.get(), self.year_combo.get(), self.semester_combo.get(),
                        self.section_combo.get(), self.student_Id_entry.get(), self.student_Name_entry.get(),
                        self.phone_entry.get(), self.email_entry.get()]

        wb = load_workbook("student_data.xlsx")
        ws = wb.active

        for row in ws.iter_rows():
            if row[5].value == updated_data[5]:  # Match Student ID
                for index, value in enumerate(updated_data):
                    ws.cell(row=row[0].row, column=index + 1, value=value)
                break

        wb.save("student_data.xlsx")
        wb.close()
        messagebox.showinfo("Success", "Record updated successfully.")
        self.show_data()  # Refresh the displayed data

    def reset_fields(self):
        self.dep_combo.set('Select Department')
        self.course_combo.set('Select Course')
        self.year_combo.set('Select Year')
        self.semester_combo.set('Select Semester')
        self.section_combo.set('Select Section')
        self.student_Id_entry.delete(0, END)
        self.student_Name_entry.delete(0, END)
        self.phone_entry.delete(0, END)
        self.email_entry.delete(0, END)

    def search_data(self):
        search_by = self.search_combo.get()
        search_text = self.search_entry.get().lower()

        # Clear the Treeview
        self.student_table.delete(*self.student_table.get_children())

        # Read data from Excel file
        wb = load_workbook("student_data.xlsx")
        ws = wb.active

        # Iterate through rows in the Excel sheet
        for row in ws.iter_rows(values_only=True):
            # Check if the search text matches the desired field (Student ID or Student Name)
            if (search_by == "Student ID" and search_text in str(row[5]).lower()) or \
            (search_by == "Student Name" and search_text in str(row[6]).lower()):
                # Insert the matched data into the Treeview widget
                self.student_table.insert('', 'end', values=row)

        wb.close()

    def show_data(self):
        # Clear existing data in the Treeview widget
        for item in self.student_table.get_children():
            self.student_table.delete(item)

        # Read data from Excel file
        wb = load_workbook("student_data.xlsx")
        ws = wb.active

        # Iterate through rows in the Excel sheet
        for row in ws.iter_rows(min_row = 2, values_only=True):
            # Insert the new data into the Treeview widget
            self.student_table.insert('', 'end', values=row)

        wb.close()

    def capture_photo(self, scholar_id, student_name, message):
        # Initialize webcam capture
        cap = cv2.VideoCapture(0)
        if not cap.isOpened():
            messagebox.showerror("Error", "Unable to access webcam.")
            return

        # Display camera feed
        while True:
            ret, frame = cap.read()
            if not ret:
                break
            cv2.imshow('Camera Feed', frame)
            key = cv2.waitKey(1) & 0xFF
            if key == ord('q'):
                break
            elif key == ord(' '):  # Capture photo when spacebar is pressed
                img_name = f"Faces/{scholar_id}_{student_name}.jpg"
                cv2.imwrite(img_name, frame)
                messagebox.showinfo("Success", message)
                break

        # Release capture resources
        cap.release()
        cv2.destroyAllWindows()

    def take_photo(self):
        # Check if a student is selected
        selected_item = self.student_table.selection()
        if not selected_item:
            messagebox.showerror("Error", "Please select a student.")
            return

        # Retrieve scholar ID and student name from the selected student's data
        scholar_id = self.student_table.item(selected_item, 'values')[5]
        student_name = self.student_table.item(selected_item, 'values')[6]

        # Capture photo
        self.capture_photo(scholar_id, student_name, "Photo captured successfully.")

    def update_photo(self):
        # Check if a student is selected
        selected_item = self.student_table.selection()
        if not selected_item:
            messagebox.showerror("Error", "Please select a student.")
            return

        # Retrieve scholar ID and student name from the selected student's data
        scholar_id = self.student_table.item(selected_item, 'values')[5]
        student_name = self.student_table.item(selected_item, 'values')[6]

        # Capture photo
        self.capture_photo(scholar_id, student_name, "Photo updated successfully.")


if __name__ == "__main__":
    root = Tk()
    app = RegistrationPage(root)
    root.mainloop()
