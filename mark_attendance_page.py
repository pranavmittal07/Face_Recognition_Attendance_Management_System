from tkinter import *
from tkinter import messagebox
from PIL import Image, ImageTk
import cv2
import face_recognition
import os
from datetime import datetime
from tkinter import ttk
import tkinter as tk
from openpyxl import load_workbook, Workbook
import shutil

class MarkAttendancePage:
    def __init__(self, root):
        self.root = root
        self.root.geometry("1530x790+0+0")
        self.root.title("Attendance")

        # Set the background color
        self.root.configure(bg='light blue')

        # top image
        img = Image.open(r"Images\cover.jpeg")
        img = img.resize((1530, 130))
        self.top = ImageTk.PhotoImage(img)

        top = Label(self.root, image=self.top)
        top.place(x=0, y=0, width=1530, height=130)

        # Title
        title_lbl = Label(self.root, text="FACE RECOGNITION ATTENDANCE SYSTEM", font=("times new roman", 35, "bold"), bg="white", fg="black")
        title_lbl.place(x=0, y=130, width=1530, height=45)

        # main frame
        main_frame = Frame(self.root, bd=2, bg="white")
        main_frame.place(x=20, y=180, width=1480, height=600)

        # left label frame
        Left_frame = LabelFrame(main_frame, bd=2, bg="white", relief=RIDGE, text="Student Details", font=("times new roman", 12, "bold"))
        Left_frame.place(x=10, y=10, width=730, height=580)

        img_left = Image.open(r"Images\20210121collegestudents.jpg")
        img_left = img_left.resize((720, 130))
        self.photoimg_left = ImageTk.PhotoImage(img_left)
        img_left_lbl = Label(Left_frame, image=self.photoimg_left)
        img_left_lbl.place(x=5, y=0, width=720, height=130)

        # current Course
        current_course_frame = LabelFrame(Left_frame, bd=2, bg="white", relief=RIDGE, text="Current Course", font=("times new roman", 12, "bold"))
        current_course_frame.place(x=5, y=140, width=720, height=300)

        # Department
        dep_lbl = Label(current_course_frame, text="Department", font=("times new roman", 12, "bold"), bg="white")
        dep_lbl.grid(row=0, column=0, padx=10, sticky=W)

        dep_combo = ttk.Combobox(current_course_frame, font=("times new roman", 10, "bold"), state="readonly", width=17)
        dep_combo["values"] = ("Select Department", "CSE", "ECE", "Electrical", "Mechanical", "MDS", "Chemical", "MME")
        dep_combo.current(0)
        dep_combo.grid(row=0, column=1, padx=2, pady=10, sticky=W)
        self.dep_combo = dep_combo

        # Course
        course_lbl = Label(current_course_frame, text="Course", font=("times new roman", 12, "bold"), bg="white")
        course_lbl.grid(row=0, column=2, padx=10, sticky=W)

        course_combo = ttk.Combobox(current_course_frame, font=("times new roman", 10, "bold"), state="readonly", width=17)
        course_combo["values"] = ("Select Course", "B.Tech", "M.Tech", "PHD")
        course_combo.current(0)
        course_combo.grid(row=0, column=3, padx=2, pady=10, sticky=W)
        self.course_combo = course_combo

        # Year
        year_lbl = Label(current_course_frame, text="Year", font=("times new roman", 12, "bold"), bg="white")
        year_lbl.grid(row=1, column=0, padx=10, sticky=W)
        year_combo = ttk.Combobox(current_course_frame, font=("times new roman", 10, "bold"), state="readonly", width=17)
        year_combo["values"] = ("Select Year", "1st Year", "2nd Year", "3rd Year", "4th Year")
        year_combo.current(0)
        year_combo.grid(row=1, column=1, padx=2, pady=10, sticky=W)
        self.year_combo = year_combo

        # Semester
        semester_lbl = Label(current_course_frame, text="Semester", font=("times new roman", 12, "bold"), bg="white")
        semester_lbl.grid(row=1, column=2, padx=10, sticky=W)
        semester_combo = ttk.Combobox(current_course_frame, font=("times new roman", 10, "bold"), state="readonly", width=17)
        semester_combo["values"] = ("Select Semester", "1st Semester", "2nd Semester", "3rd Semester", "4th Semester", "5th Semester", "6th Semester", "7th Semester", "8th Semester")
        semester_combo.current(0)
        semester_combo.grid(row=1, column=3, padx=2, pady=10, sticky=W)
        self.semester_combo = semester_combo

        # section
        section_lbl = Label(current_course_frame, text="Section", font=("times new roman", 12, "bold"), bg="white")
        section_lbl.grid(row=2, column=0, padx=10, pady=5, sticky=W)
        section_combo = ttk.Combobox(current_course_frame, font=("times new roman", 10, "bold"), state="readonly", width=17)
        section_combo["values"] = ("Select Section", "I", "II", "III")
        section_combo.current(0)
        section_combo.grid(row=2, column=1, padx=2, pady=5, sticky=W)
        self.section_combo = section_combo

        # subject
        subject_lbl = Label(current_course_frame, text="Subject", font=("times new roman", 12, "bold"), bg="white")
        subject_lbl.grid(row=3, column=0, padx=10, pady=5, sticky=W)
        self.subject_entry = ttk.Entry(current_course_frame, width=20, font=("times new roman", 10, "bold"))
        self.subject_entry.grid(row=3, column=1, padx=10, pady=5, sticky=W)

        # Teacher Name
        teacher_lbl = Label(current_course_frame, text="Teacher Name", font=("times new roman", 12, "bold"), bg="white")
        teacher_lbl.grid(row=4, column=0, padx=10, pady=5, sticky=W)
        self.teacher_entry = ttk.Entry(current_course_frame, width=20, font=("times new roman", 10, "bold"))
        self.teacher_entry.grid(row=4, column=1, padx=10, pady=5, sticky=W)

        # Enter Button
        self.Enter_rand_btn = Button(current_course_frame, text="Enter", width=35, font=("times new roman", 13, "bold"), bg="blue", fg="white", command=self.Enter)
        self.Enter_rand_btn.place(x=250, y=200, width=150, height=40)

        #self.Enter_btn = Button(current_course_frame, text="Enter [In class]", width=35, font=("times new roman", 13, "bold"), bg="blue", fg="white", command=self.Enter)
        #self.Enter_btn.place(x=300, y=200, width=150, height=40)

        # Camera frame
        self.camera_frame = LabelFrame(main_frame, bd=2, bg="white", relief=RIDGE, text="Camera", font=("times new roman", 12, "bold"))
        self.camera_frame.place(x=750, y=10, width=720, height=580)

        self.canvas = Canvas(self.camera_frame, bg="black")
        self.canvas.pack(fill=BOTH, expand=True)

        # Capture image button & Download
        btn_frame1 = Frame(Left_frame, bd=2, relief=RIDGE, bg="light pink")
        btn_frame1.place(x=0, y=450, width=730, height=80)

        capture_btn = Button(btn_frame1, text="Take Attendance", width=36, font=("times new roman", 13, "bold"), bg="blue", fg="white", command=self.capture_and_recognize)
        capture_btn.grid(row=1, column=0)

        download_btn = Button(btn_frame1, text="Download Attendance", width=36, font=("times new roman", 13, "bold"), bg="blue", fg="white", command=self.download)
        download_btn.grid(row=1, column=1)

        # Directory containing known face images
        self.known_faces_dir = 'Faces'  # You might need to adjust the directory path

        # Load known faces
        self.known_faces = []
        self.known_names = []
        self.scholar_numbers = []
        for filename in os.listdir(self.known_faces_dir):
            if filename.endswith('.jpg') or filename.endswith('.png'):
                # Parse the scholar number and the name from the filename
                scholar_number, name = filename.split('_')
                image = face_recognition.load_image_file(os.path.join(self.known_faces_dir, filename))
                encoding = face_recognition.face_encodings(image)[0]
                self.known_faces.append(encoding)
                self.known_names.append(name)
                self.scholar_numbers.append(scholar_number)

    def capture_and_recognize(self):
        # Capture image
        captured_image = self.capture_image()
        if captured_image is None:
            return
        # Recognize face
        student_name, scholar_number = self.recognize_face(captured_image)
        if student_name is None:
            messagebox.showinfo("Recognition Error", "Student not recognized!")
            return
        # Mark attendance
        self.mark_attendance(scholar_number, student_name)

    def capture_image(self):
        cam = cv2.VideoCapture(0)
        while True:
            ret, frame = cam.read()
            if not ret:
                print("Failed to grab frame")
                messagebox.showerror("Capture Error", "Failed to grab frame")
                return None
                break
            cv2.imshow('Press Space to capture', frame)
            if cv2.waitKey(1) & 0xFF == ord(' '):
                break
        cam.release()
        cv2.destroyAllWindows()
        return frame if ret else None
    
    def recognize_face(self, captured_image):
        face_encodings = face_recognition.face_encodings(captured_image)
        if len(face_encodings) == 0:
            messagebox.showinfo("Face Detection", "No faces detected in the image.")
            return None, None
        captured_encoding = face_encodings[0]
        matches = face_recognition.compare_faces(self.known_faces, captured_encoding)
        if True in matches:
            first_match_index = matches.index(True)
            return self.known_names[first_match_index], self.scholar_numbers[first_match_index]
        return None, None

    def Enter(self):
        # File path to fetch from the "Classes" folder
        file_name = f"{self.dep_combo.get()}_{self.year_combo.get()}_{self.section_combo.get()}.xlsx"
        file_path = os.path.join("Classes", file_name)

        if os.path.exists(file_path):
            # Check if the new file already exists in the "Attendance" folder
            new_file_name = f"{self.dep_combo.get()}_{self.year_combo.get()}_{self.section_combo.get()}_{self.subject_entry.get()}.xlsx"
            new_file_path = os.path.join("Attendance", new_file_name)

            if os.path.exists(new_file_path):
                # File exists, load workbook and make changes
                wb = load_workbook(new_file_path)
                ws = wb.active
            else:
                # File does not exist, create new workbook and copy data
                wb = Workbook()
                ws = wb.active

                # Load existing workbook directly into the new workbook
                existing_wb = load_workbook(file_path)
                existing_ws = existing_wb.active

                for row in existing_ws.iter_rows():
                    for cell in row:
                        ws[cell.coordinate].value = cell.value

                # Save the new workbook directly to the new file path
                wb.save(new_file_path)

            # Make changes in the worksheet
            row_sub = 13
            col_sub = 2
            row_teach = 14
            col_teach = 2
            ws.cell(row=row_sub, column=col_sub, value=self.subject_entry.get())  
            ws.cell(row=row_teach, column=col_teach, value=self.teacher_entry.get())

            # Find the current date
            current_date = datetime.now().strftime("%Y-%m-%d")

            # Increment the column index when marking attendance
            column_index = 3  # Start from column 3
            while ws.cell(row=17, column=column_index).value is not None:
                column_index += 1

            # Mark the current date in row 17 and the incremented column
            ws.cell(row=17, column=column_index, value=current_date)

            # Save the workbook
            wb.save(new_file_path)
            messagebox.showinfo("File Saved", "Attendance details saved successfully.")
        else:
            messagebox.showerror("File Not Found", "File not found in the Classes folder.")

    def download(self):
        # File name to download
        file_name = f"{self.dep_combo.get()}_{self.year_combo.get()}_{self.section_combo.get()}_{self.subject_entry.get()}.xlsx"
        file_path = os.path.join("Attendance", file_name)

        # Check if the file exists in the "Attendance" folder
        if os.path.exists(file_path):
            try:
                # Get the user's downloads directory
                downloads_dir = os.path.expanduser('~')
                downloads_dir = os.path.join(downloads_dir, "Downloads")

                # Destination path to save the file in the downloads directory
                destination_path = os.path.join(downloads_dir, file_name)

                # Copy the file to the downloads directory
                shutil.copyfile(file_path, destination_path)

                messagebox.showinfo("Download Complete", f"{file_name} downloaded successfully to Downloads folder.")
            except Exception as e:
                messagebox.showerror("Download Error", f"An error occurred while downloading the file: {str(e)}")
        else:
            messagebox.showerror("File Not Found", f"{file_name} not found in the Attendance folder.")

    def mark_attendance(self, scholar_number, student_name):
        # Construct file path
        file_name = f"{self.dep_combo.get()}_{self.year_combo.get()}_{self.section_combo.get()}_{self.subject_entry.get()}.xlsx"
        file_path = os.path.join("Attendance", file_name)

        if os.path.exists(file_path):
            now = datetime.now()
            current_time = now.strftime("%H:%M:%S")

            wb = load_workbook(file_path)
            ws = wb.active

            start_row = 18  # Start searching from row 18
            found = False
            present_column = None

            # Search for the scholar number in column 1 from row 18 to the last filled row
            for row in range(start_row, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=1).value
                if str(cell_value).strip() == str(scholar_number).strip():  # Ensure both are treated as strings
                    found = True
                    present_column = ws.max_column  # Get the last filled column of row 17
                    break

            if found:
                # Mark "P" in the identified column for the scholar number row
                ws.cell(row=row, column=present_column, value="P")

                # Save the workbook
                wb.save(file_path)
                messagebox.showinfo("Attendance Marked", f"Attendance marked for {student_name} at {current_time}.")
            else:
                messagebox.showerror("Scholar Number Not Found", f"Scholar number {scholar_number} not found in the Excel file.")
        else:
            messagebox.showerror("File Not Found", f"File '{file_name}' not found in the Attendance folder.")

# Main execution
if __name__ == "__main__":
    root = Tk()
    app = MarkAttendancePage(root)
    root.mainloop()
